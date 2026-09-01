"""
Excel Exporter Module
Handles creation and formatting of Excel output files
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
from config import OUTPUT_DIR, OUTPUT_FILENAME_PREFIX, EXCEL_HEADERS


class ExcelExporter:
    """Handles Excel file creation and formatting"""
    
    def __init__(self):
        """Initialize the Excel exporter"""
        # Create output directory if it doesn't exist
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)
    
    def export_to_excel(self, data, keyword=None):
        """
        Export scraped data to Excel file with all data in a single sheet
        
        Args:
            data (list): List of dictionaries containing product data
            keyword (str): Optional keyword to include in filename
            
        Returns:
            str: Path to the created Excel file
        """
        if not data:
            print("No data to export!")
            return None
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        keyword_part = f"_{keyword}" if keyword else ""
        filename = f"{OUTPUT_FILENAME_PREFIX}{keyword_part}_{timestamp}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        # Clean price values to extract only numbers
        import re
        cleaned_data = []
        for row in data:
            cleaned_row = row.copy()
            if 'Price' in cleaned_row and cleaned_row['Price']:
                # Extract numeric value from price string (e.g., "AED 1942.45" -> 1942.45)
                price_str = str(cleaned_row['Price'])
                # Remove currency codes and extract the full decimal number
                # Pattern matches: optional digits, optional decimal point, optional digits
                match = re.search(r'(\d+(?:\.\d+)?)', price_str)
                if match:
                    try:
                        cleaned_row['Price'] = float(match.group(1))
                    except:
                        pass  # Keep original if conversion fails
            cleaned_data.append(cleaned_row)
        
        # Create DataFrame with all data in a single sheet
        df = pd.DataFrame(cleaned_data, columns=EXCEL_HEADERS)
        
        # Create Excel writer and export to single sheet
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='All Products', index=False)
        
        # Format the Excel file
        self._format_excel(filepath)
        
        # Print summary
        print(f"\nExcel file created: {filepath}")
        print(f"Sheet: All Products")
        print(f"Total rows: {len(data)}")
        
        # Show breakdown by region if Region column exists
        if 'Region' in data[0]:
            region_counts = {}
            for row in data:
                region = row.get('Region', 'Unknown')
                region_counts[region] = region_counts.get(region, 0) + 1
            
            print("\nBreakdown by region:")
            for region, count in sorted(region_counts.items()):
                print(f"  - {region}: {count} rows")
        
        return filepath
    
    def _format_excel(self, filepath):
        """
        Apply formatting to all sheets in Excel file
        
        Args:
            filepath (str): Path to Excel file
        """
        # Load workbook
        wb = load_workbook(filepath)
        
        # Format each sheet
        for ws in wb.worksheets:
            # Header formatting
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set width with some padding
                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
        
        # Save formatted workbook
        wb.save(filepath)
    
    def append_to_excel(self, filepath, new_data):
        """
        Append new data to existing Excel file
        
        Args:
            filepath (str): Path to existing Excel file
            new_data (list): List of dictionaries to append
        """
        # Read existing data
        df_existing = pd.read_excel(filepath)
        
        # Create DataFrame from new data
        df_new = pd.DataFrame(new_data, columns=EXCEL_HEADERS)
        
        # Combine
        df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        
        # Write back
        df_combined.to_excel(filepath, index=False, sheet_name='Products')
        
        # Re-format
        self._format_excel(filepath)
        
        print(f"Appended {len(new_data)} rows to {filepath}")
